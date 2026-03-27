"""
원본 엑셀 데이터 → 학습용 정제 데이터 변환 스크립트

처리 항목:
1. 결측값/빈 텍스트 제거
2. 중복 제거
3. 라벨 통일 (긍정/positive/POSITIVE → 1, 부정/negative/NEGATIVE → 0)
4. 라벨 누락 제거
5. 텍스트 정제 (HTML 태그, URL, 이모지, 과도한 반복문자, 스팸 제거)
6. 길이 이상치 필터링
7. 날짜 형식 통일
8. train/eval 분할 후 JSON 저장
"""

import json
import re
import random
from pathlib import Path

import openpyxl

# ──────────────────────────────────────────
# 설정
# ──────────────────────────────────────────
RAW_PATH = Path(__file__).parent / "raw_reviews.xlsx"
OUTPUT_DIR = Path(__file__).parent
MIN_TEXT_LENGTH = 5          # 최소 텍스트 길이
MAX_TEXT_LENGTH = 300        # 최대 텍스트 길이
EVAL_RATIO = 0.2             # 검증 데이터 비율
SEED = 42

# 라벨 매핑
LABEL_MAP = {
    "긍정": 1, "positive": 1, "POSITIVE": 1, "pos": 1,
    "부정": 0, "negative": 0, "NEGATIVE": 0, "neg": 0,
}


# ──────────────────────────────────────────
# 전처리 함수들
# ──────────────────────────────────────────
def load_excel(path: Path) -> list[dict]:
    """엑셀 파일 로드 → dict 리스트 변환"""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    print(f"[1/7] 엑셀 로드 완료: {len(rows)}건")
    return rows


def remove_missing(data: list[dict]) -> list[dict]:
    """텍스트 또는 라벨이 없거나 빈 문자열인 행 제거"""
    before = len(data)
    result = []
    for row in data:
        text = row.get("text")
        label = row.get("label")
        if text is None or label is None:
            continue
        if isinstance(text, str) and text.strip() == "":
            continue
        result.append(row)
    removed = before - len(result)
    print(f"[2/7] 결측값 제거: {removed}건 제거 → {len(result)}건 남음")
    return result


def unify_labels(data: list[dict]) -> list[dict]:
    """라벨을 0(부정), 1(긍정)으로 통일. 매핑 불가 라벨은 제거"""
    before = len(data)
    result = []
    unknown_labels = set()
    for row in data:
        raw_label = str(row["label"]).strip()
        mapped = LABEL_MAP.get(raw_label)
        if mapped is None:
            unknown_labels.add(raw_label)
            continue
        row["label"] = mapped
        result.append(row)
    removed = before - len(result)
    if unknown_labels:
        print(f"  ⚠ 매핑 불가 라벨: {unknown_labels}")
    print(f"[3/7] 라벨 통일: {removed}건 제거 → {len(result)}건 남음")
    return result


def clean_text(text: str) -> str:
    """텍스트 정제"""
    # HTML 태그 제거
    text = re.sub(r"<[^>]+>", "", text)
    # URL 제거
    text = re.sub(r"https?://\S+", "", text)
    # 이모지 제거
    text = re.sub(
        r"[\U0001F600-\U0001F64F\U0001F300-\U0001F5FF"
        r"\U0001F680-\U0001F6FF\U0001F900-\U0001F9FF"
        r"\U00002702-\U000027B0\U00002600-\U000026FF"
        r"\u2764\uFE0F\u2665\u200d]+",
        "", text
    )
    # 반복 자모 축소 (ㅋㅋㅋㅋㅋ → ㅋㅋ, ㅠㅠㅠㅠ → ㅠㅠ)
    text = re.sub(r"([ㄱ-ㅎㅏ-ㅣ])\1{2,}", r"\1\1", text)
    # 반복 특수문자 축소 (!!!!!! → !)
    text = re.sub(r"([!?.])\1{2,}", r"\1", text)
    # 앞뒤 공백 정리
    text = text.strip()
    return text


def apply_text_cleaning(data: list[dict]) -> list[dict]:
    """텍스트 정제 적용"""
    for row in data:
        row["text"] = clean_text(str(row["text"]))
    print(f"[4/7] 텍스트 정제 완료: {len(data)}건")
    return data


def remove_spam_and_noise(data: list[dict]) -> list[dict]:
    """스팸, 의미없는 텍스트 필터링"""
    before = len(data)
    result = []
    spam_patterns = [
        r"(?i)(무료\s*다운|클릭|광고|홍보|bit\.ly)",
        r"^[ㄱ-ㅎㅏ-ㅣa-zA-Z\s]+$",  # 자음/모음/영문만으로 구성
    ]
    for row in data:
        text = row["text"]
        is_spam = any(re.search(p, text) for p in spam_patterns)
        if is_spam:
            continue
        result.append(row)
    removed = before - len(result)
    print(f"[5/7] 스팸/노이즈 제거: {removed}건 제거 → {len(result)}건 남음")
    return result


def filter_by_length(data: list[dict]) -> list[dict]:
    """텍스트 길이 기준 필터링"""
    before = len(data)
    result = []
    for row in data:
        length = len(row["text"])
        if MIN_TEXT_LENGTH <= length <= MAX_TEXT_LENGTH:
            result.append(row)
    removed = before - len(result)
    print(f"[6/7] 길이 필터링 ({MIN_TEXT_LENGTH}~{MAX_TEXT_LENGTH}자): {removed}건 제거 → {len(result)}건 남음")
    return result


def remove_duplicates(data: list[dict]) -> list[dict]:
    """텍스트 기준 중복 제거 (첫 번째만 유지)"""
    before = len(data)
    seen = set()
    result = []
    for row in data:
        text = row["text"]
        if text in seen:
            continue
        seen.add(text)
        result.append(row)
    removed = before - len(result)
    print(f"[7/7] 중복 제거: {removed}건 제거 → {len(result)}건 남음")
    return result


def split_and_save(data: list[dict], output_dir: Path):
    """train/eval 분할 후 JSON 저장"""
    random.seed(SEED)
    random.shuffle(data)

    # 학습에 필요한 필드만 추출
    clean_data = [{"text": row["text"], "label": row["label"]} for row in data]

    split_idx = int(len(clean_data) * (1 - EVAL_RATIO))
    train_data = clean_data[:split_idx]
    eval_data = clean_data[split_idx:]

    train_path = output_dir / "train.json"
    eval_path = output_dir / "eval.json"

    with open(train_path, "w", encoding="utf-8") as f:
        json.dump(train_data, f, ensure_ascii=False, indent=2)
    with open(eval_path, "w", encoding="utf-8") as f:
        json.dump(eval_data, f, ensure_ascii=False, indent=2)

    # 라벨 분포 출력
    for name, split in [("train", train_data), ("eval", eval_data)]:
        pos = sum(1 for d in split if d["label"] == 1)
        neg = len(split) - pos
        print(f"  {name}: {len(split)}건 (긍정: {pos}, 부정: {neg})")

    print(f"\n저장 완료:")
    print(f"  → {train_path}")
    print(f"  → {eval_path}")


# ──────────────────────────────────────────
# 메인 실행
# ──────────────────────────────────────────
def main():
    print("=" * 50)
    print("전처리 시작")
    print("=" * 50)

    data = load_excel(RAW_PATH)
    data = remove_missing(data)
    data = unify_labels(data)
    data = apply_text_cleaning(data)
    data = remove_spam_and_noise(data)
    data = filter_by_length(data)
    data = remove_duplicates(data)

    print(f"\n최종 정제 데이터: {len(data)}건")
    print("-" * 50)

    split_and_save(data, OUTPUT_DIR)
    print("=" * 50)
    print("전처리 완료!")


if __name__ == "__main__":
    main()
